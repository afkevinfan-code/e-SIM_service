#!/usr/bin/env python3
"""Search K-eSIM quote spreadsheets for matching travel SIM products."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_CARD_PATH = PROJECT_DIR / "card.xlsx"
DEFAULT_ESIM_PATH = PROJECT_DIR / "eSim.xlsx"
LEGACY_CARD_PATH = Path(
    "/Users/kevinfan/Library/Mobile Documents/com~apple~CloudDocs/K-eSIM/card.xlsx"
)
LEGACY_ESIM_PATH = Path(
    "/Users/kevinfan/Library/Mobile Documents/com~apple~CloudDocs/K-eSIM/eSim.xlsx"
)
UNLIMITED_QUERY_TERMS = {
    "不限流量",
    "吃到饱",
    "无限",
    "unlimited",
    "max",
}
UNLIMITED_PRODUCT_TERMS = ("无限", "max", "unlimited")


@dataclass
class Product:
    source: str
    name: str
    unit_price: str
    quote_cost: float | None
    destinations: str
    description: str
    note: str
    activation: str = ""


@dataclass
class SearchResult:
    score: int
    product: Product
    quote_total: float | None
    quote_label: str


def resolve_data_path(explicit_path: str | None, default_path: Path, legacy_path: Path) -> Path:
    if explicit_path:
        return Path(explicit_path).expanduser().resolve()
    if default_path.exists():
        return default_path
    return legacy_path


def normalize(text: object) -> str:
    value = str(text or "").casefold().replace(" ", "")
    country_aliases = {
        "關島": "关岛",
        "希臘": "希腊",
        "義大利": "意大利",
        "克羅埃西亞": "克罗地亚",
        "克罗埃西亚": "克罗地亚",
        "斯洛維尼亞": "斯洛文尼亚",
        "斯洛维尼亚": "斯洛文尼亚",
        "荷蘭": "荷兰",
        "愛爾蘭": "爱尔兰",
        "愛沙尼亞": "爱沙尼亚",
        "冰島": "冰岛",
        "拉脫維亞": "拉脱维亚",
        "盧森堡": "卢森堡",
        "羅馬尼亞": "罗马尼亚",
        "馬耳他": "马耳他",
        "保加利亞": "保加利亚",
        "烏克蘭": "乌克兰",
        "梵蒂岡": "梵蒂冈",
        "新加坡": "新加坡",
        "紐西蘭": "新西兰",
        "紐西兰": "新西兰",
        "南韓": "韩国",
        "韓國": "韩国",
        "美國": "美国",
        "英國": "英国",
        "德國": "德国",
        "法國": "法国",
        "泰國": "泰国",
        "中國": "中国",
        "香港": "香港",
        "澳門": "澳门",
        "阿聯酋": "阿联酋",
        "沙烏地阿拉伯": "沙特阿拉伯",
    }
    for old, new in country_aliases.items():
        value = value.replace(old, new)
    replacements = {
        "國": "国",
        "臺": "台",
        "灣": "湾",
        "韓": "韩",
        "馬": "马",
        "來": "来",
        "亞": "亚",
        "歐": "欧",
        "紐": "纽",
        "蘭": "兰",
        "羅": "罗",
        "無": "无",
        "線": "线",
        "總": "总",
        "區": "区",
        "點": "点",
        "達": "达",
        "機": "机",
        "啟": "启",
        "實": "实",
        "體": "体",
        "虛": "虚",
        "擬": "拟",
        "關": "关",
        "島": "岛",
        "爾": "尔",
        "愛": "爱",
        "脫": "脱",
        "盧": "卢",
        "烏": "乌",
        "岡": "冈",
        "聯": "联",
        "酋": "酋",
        "飽": "饱",
        "維": "维",
        "臘": "腊",
        "麗": "丽",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value


def keyword_variants(keyword: str) -> tuple[str, ...]:
    normalized = normalize(keyword)
    if normalized in UNLIMITED_QUERY_TERMS:
        return UNLIMITED_PRODUCT_TERMS
    return (normalized,)


def load_products(path: Path, source: str) -> list[Product]:
    if not path.exists():
        raise FileNotFoundError(f"{source} data file not found: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "") for cell in next(sheet.iter_rows(max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    required_headers = ["商品名称", "单价(元)", "報價成本", "使用地", "商品描述"]
    missing_headers = [name for name in required_headers if name not in index]
    if missing_headers:
        raise ValueError(
            f"{source} data file is missing required columns: {', '.join(missing_headers)}"
        )

    products: list[Product] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        products.append(
            Product(
                source=source,
                name=str(row[index["商品名称"]] or ""),
                unit_price=str(row[index["单价(元)"]] or ""),
                quote_cost=row[index["報價成本"]],
                destinations=str(row[index["使用地"]] or ""),
                description=str(row[index["商品描述"]] or ""),
                note=str(row[index.get("备注", -1)] or "") if "备注" in index else "",
                activation=str(row[index.get("激活方式", -1)] or "")
                if "激活方式" in index
                else "",
            )
        )
    return products


def score(product: Product, keywords: Iterable[str]) -> int:
    haystack = normalize(
        " ".join(
            [
                product.name,
                product.destinations,
                product.description,
                product.note,
                product.activation,
            ]
        )
    )
    value = 0
    normalized_destinations = normalize(product.destinations)
    normalized_name = normalize(product.name)
    for keyword in keywords:
        variants = keyword_variants(keyword)
        if not any(variants):
            continue
        if any(needle in normalized_destinations for needle in variants):
            value += 6
        if any(needle in normalized_name for needle in variants):
            value += 4
        if any(needle in haystack for needle in variants):
            value += 2
    return value


def matches_all(product: Product, keywords: Iterable[str]) -> bool:
    haystack = normalize(
        " ".join(
            [
                product.name,
                product.destinations,
                product.description,
                product.note,
                product.activation,
                product.source,
            ]
        )
    )
    return all(
        any(variant in haystack for variant in keyword_variants(keyword))
        for keyword in keywords
        if normalize(keyword)
    )


def matches_destinations(product: Product, destinations: Iterable[str]) -> bool:
    normalized_destinations = normalize(product.destinations)
    return all(
        normalize(destination) in normalized_destinations
        for destination in destinations
        if normalize(destination)
    )


def extract_day_hint(text: str) -> int | None:
    match = re.search(r"(\d+)\s*(?:天|日|day|days)", text, re.IGNORECASE)
    return int(match.group(1)) if match else None


def fixed_product_days(product: Product) -> int | None:
    text = f"{product.name} {product.description}"
    patterns = [
        r"[-－]\s*0*(\d+)\s*(?:天|日|day|days)",
        r"(\d+)\s*day\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def is_daily_product(product: Product) -> bool:
    text = f"{product.name} {product.description}"
    return "/天" in text or "每天" in text or "无限/天" in text or "無限/天" in text


def estimated_quote(product: Product, days: int | None) -> tuple[float | None, str]:
    if product.quote_cost is None:
        return None, ""
    fixed_days = fixed_product_days(product)
    if days is not None and fixed_days is None and is_daily_product(product):
        return product.quote_cost * days, f"{days} 天估算"
    return product.quote_cost, "商品報價"


def meaningful_keywords(keywords: list[str]) -> list[str]:
    return [
        keyword
        for keyword in keywords
        if not re.fullmatch(r"\d+\s*(?:天|日|day|days)", keyword, re.IGNORECASE)
    ]


def day_score(product: Product, days: int | None) -> int:
    if days is None:
        return 0
    fixed_days = fixed_product_days(product)
    if fixed_days == days:
        return 8
    if fixed_days is None and is_daily_product(product):
        return 3
    if fixed_days is None:
        return 0
    return -4


def format_money(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:,.0f}" if float(value).is_integer() else f"{value:,.2f}"


def load_all_products(
    product_type: str = "all",
    card_path: str | None = None,
    esim_path: str | None = None,
) -> list[Product]:
    resolved_card_path = resolve_data_path(card_path, DEFAULT_CARD_PATH, LEGACY_CARD_PATH)
    resolved_esim_path = resolve_data_path(esim_path, DEFAULT_ESIM_PATH, LEGACY_ESIM_PATH)

    products: list[Product] = []
    if product_type in {"all", "card"}:
        products.extend(load_products(resolved_card_path, "實體卡"))
    if product_type in {"all", "esim"}:
        products.extend(load_products(resolved_esim_path, "eSIM"))
    return products


def search_products(
    keywords: list[str],
    product_type: str = "all",
    limit: int | None = 12,
    card_path: str | None = None,
    esim_path: str | None = None,
    destination_keywords: list[str] | None = None,
    plan_keywords: list[str] | None = None,
) -> list[SearchResult]:
    destination_keywords = destination_keywords or []
    plan_keywords = plan_keywords if plan_keywords is not None else keywords
    query_text = " ".join([*destination_keywords, *plan_keywords])
    days = extract_day_hint(query_text)
    effective_plan_keywords = meaningful_keywords(plan_keywords)
    products = load_all_products(product_type, card_path, esim_path)

    ranked: list[SearchResult] = []
    for product in products:
        if destination_keywords and not matches_destinations(product, destination_keywords):
            continue
        if effective_plan_keywords and not matches_all(product, effective_plan_keywords):
            continue
        keyword_score = score(product, destination_keywords) + score(
            product, effective_plan_keywords
        )
        if keyword_score <= 0 and (destination_keywords or effective_plan_keywords):
            continue
        quote_total, quote_label = estimated_quote(product, days)
        ranked.append(
            SearchResult(
                score=keyword_score + day_score(product, days),
                product=product,
                quote_total=quote_total,
                quote_label=quote_label,
            )
        )

    ranked.sort(
        key=lambda item: (
            item.quote_total if item.quote_total is not None else float("inf"),
            -item.score,
        )
    )
    return ranked[:limit] if limit is not None else ranked


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Search physical SIM and eSIM quote spreadsheets."
    )
    parser.add_argument("keywords", nargs="+", help="Destination and plan keywords")
    parser.add_argument("--limit", type=int, default=12, help="Maximum results")
    parser.add_argument("--type", choices=["all", "card", "esim"], default="all")
    parser.add_argument(
        "--card-path",
        help="Physical SIM workbook path. Defaults to ./card.xlsx, then legacy iCloud path.",
    )
    parser.add_argument(
        "--esim-path",
        help="eSIM workbook path. Defaults to ./eSim.xlsx, then legacy iCloud path.",
    )
    args = parser.parse_args()

    try:
        ranked = search_products(
            args.keywords,
            product_type=args.type,
            limit=args.limit,
            card_path=args.card_path,
            esim_path=args.esim_path,
        )
    except (FileNotFoundError, ValueError) as error:
        print(f"資料檔讀取失敗：{error}", file=sys.stderr)
        raise SystemExit(1) from error

    if not ranked:
        print("找不到符合條件的商品。請換目的地名稱、英文國名，或放寬天數/流量條件。")
        return

    query_text = " ".join(args.keywords)
    print(f"查詢：{query_text}")
    print(f"結果：前 {min(args.limit, len(ranked))} 筆候選商品")
    print()
    for position, result in enumerate(ranked, start=1):
        product = result.product
        print(f"{position}. [{product.source}] {product.name}")
        print(
            f"   {result.quote_label}：{format_money(result.quote_total)}"
            f"｜原報價成本：{format_money(product.quote_cost)}"
            f"｜单价(元)：{product.unit_price}"
        )
        print(f"   使用地：{product.destinations}")
        print(f"   內容：{product.description}")
        if product.note:
            print(f"   備註：{product.note}")
        if product.activation:
            print(f"   激活：{product.activation}")
        print()


if __name__ == "__main__":
    main()
